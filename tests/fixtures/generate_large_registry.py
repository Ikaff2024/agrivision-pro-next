"""Generateur de registre cooperative SYNTHETIQUE a grande echelle.

Remplace le registre client reel qui servait auparavant de fixture. Le fichier
produit reproduit la STRUCTURE attendue par
``app.importers.cooperative_registry`` — pas les donnees d'une cooperative.

Ce que le test doit continuer de couvrir, et que ce generateur preserve :

  * le VOLUME (~7 000 producteurs / ~7 100 parcelles), pour que le parseur reste
    eprouve a l'echelle reelle et que les regressions de performance restent
    visibles ;
  * la POSITION des en-tetes : lignes 5-6 pour la feuille Producteurs, 5-7 pour
    la feuille Plantations, donnees a partir de la ligne 8 ;
  * la DETECTION des colonnes par mots-cles (code, nom, localite, contact,
    piece d'identification, superficie, GPS...) ;
  * les COLONNES AMBIGUES : « sexe » et « date de naissance » apparaissent deux
    fois (menage puis producteur) et le parseur doit retenir la DERNIERE ;
  * les CELLULES FUSIONNEES : sur la feuille Plantations, le code producteur
    n'est ecrit que sur la premiere ligne d'un groupe et doit etre propage ;
  * la feuille FORMATIONS en matrice (codes en colonne A, thematiques en ligne) ;
  * la DETECTION DE CAMPAGNE depuis le nom de fichier.

Toutes les valeurs sont manifestement fictives et deterministes (seed fixe) :
aucun nom, contact, piece d'identite, localite ou coordonnee ne provient d'un
fichier client. Les telephones utilisent le prefixe de documentation +99, les
coordonnees sont posees sur une grille arbitraire.

Usage :
    from tests.fixtures.generate_large_registry import build_large_registry
    path = build_large_registry(tmp_path)          # genere a la volee

Le XLSX n'est JAMAIS versionne : il est ecrit dans le tmp_path de pytest, ce qui
evite d'ajouter plusieurs Mo de binaire au depot.
"""
from __future__ import annotations

import random
from pathlib import Path

from openpyxl import Workbook

# Convention de code volontairement neutre : aucun lien avec une cooperative.
CODE_PREFIX = "COOPDEMO"
CAMPAIGN = "2025-2026"
FILENAME = f"Registre DEMO SYNTHETIQUE {CAMPAIGN} Act.xlsx"

# Echelle : au-dessus des seuils historiques (>6900 producteurs, >7000 parcelles)
# pour que le test de volume garde exactement la meme signification.
N_PRODUCERS = 7_050
N_EXTRA_PLANTATIONS = 120      # producteurs possedant une 2e parcelle
N_FORMATION_ROWS = 900

_SECTIONS = [f"SECTION_TEST_{i:02d}" for i in range(1, 13)]
_LOCALITES = [f"LOCALITE_TEST_{i:03d}" for i in range(1, 61)]
_PIECE_NATURES = ["CNI_TEST", "ATTESTATION_TEST", "PASSEPORT_TEST"]
_PROJETS = ["FT-RA", "FT", "RA", "EUDR"]
_THEMES = [
    "Travail des enfants", "Bonnes pratiques agricoles", "Agroforesterie",
    "Sante et securite", "Egalite de genre", "Deforestation",
]

# En-tetes de la feuille Producteurs. L'ordre reproduit le registre reel :
# le bloc « membre du menage » (sexe / date de naissance) PRECEDE celui du
# producteur, si bien que le parseur doit retenir la DERNIERE occurrence.
PRODUCER_HEADERS = [
    "N°",                          # 1
    "CODE YEYA",                   # 2  -> code producteur (mots-cles code+yeya)
    "Nom et Prenoms des Membres",  # 3  -> nom (code+nom+membre)
    "Membre du menage",            # 4
    "SEXE",                        # 5  <- occurrence MENAGE (leurre)
    "DATE DE NAISSANCE",           # 6  <- occurrence MENAGE (leurre)
    "Qui travaille sur la plantation",  # 7
    "SEXE",                        # 8  <- occurrence PRODUCTEUR (retenue)
    "Date de Naissance",           # 9  <- occurrence PRODUCTEUR (retenue)
    "CODE PLANTATION",             # 10
    "CODE SACO",                   # 11
    "RECEPISSE",                   # 12
    "PROJET",                      # 13
    "Nom du Formateur Interne",    # 14
    "Section",                     # 15
    "Localites",                   # 16
    "Contacts",                    # 17
    "N° de la piece d'identification",  # 18
    "Nature de la piece",          # 19
    "LATITUDE",                    # 20
    "LONGITUDE",                   # 21
]

PLANTATION_HEADERS = [
    "CODE PRODUCTEUR",             # 1
    "CODE PLANTATION",             # 2
    "PROJET",                      # 3
    "Superficie de la Plantation (ha)",  # 4
    "LATITUDE",                    # 5
    "LONGITUDE",                   # 6
    "Rendement 2025-2026",         # 7
    "Production Estimee 2025-2026",      # 8
]


def producer_code(index: int) -> str:
    """Code producteur synthetique, stable et trivialement reconnaissable."""
    return f"{CODE_PREFIX}{index:05d}"


def _fake_phone(rng: random.Random) -> str:
    # +99 : prefixe pays reserve/inexistant -> aucun numero joignable.
    return f"+99{rng.randint(10**8, 10**9 - 1)}"


def _write_headers(ws, headers: list[str], row: int) -> None:
    for col, label in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=label)


def _build_producers_sheet(wb: Workbook, rng: random.Random) -> None:
    ws = wb.create_sheet("Registre Producteurs")
    ws.cell(row=1, column=1, value=f"COOP_DEMO — REGISTRE DES PRODUCTEURS {CAMPAIGN}")
    # En-tetes lignes 5-6 (le parseur lit build_headers(ws, [5, 6])).
    _write_headers(ws, PRODUCER_HEADERS, 5)

    row = 8  # la ligne 7 reste vide, comme dans le registre d'origine
    for i in range(1, N_PRODUCERS + 1):
        code = producer_code(i)
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=code)
        ws.cell(row=row, column=3, value=f"PRODUCTEUR_TEST_{i:05d}")
        ws.cell(row=row, column=4, value=f"MENAGE_TEST_{i:05d}")
        # Bloc menage : valeurs volontairement DIFFERENTES du bloc producteur,
        # pour qu'un parseur qui lirait la mauvaise occurrence soit demasque.
        ws.cell(row=row, column=5, value="F")
        ws.cell(row=row, column=6, value="1990-01-01")
        ws.cell(row=row, column=7, value="Oui" if i % 3 == 0 else "Non")
        ws.cell(row=row, column=8, value="M")
        ws.cell(row=row, column=9, value="1975-06-15")
        ws.cell(row=row, column=10, value=f"{code}-P1")
        ws.cell(row=row, column=11, value=f"SACO_TEST_{i:05d}")
        ws.cell(row=row, column=12, value=f"REC_TEST_{i:05d}")
        ws.cell(row=row, column=13, value=_PROJETS[i % len(_PROJETS)])
        ws.cell(row=row, column=14, value=f"FORMATEUR_TEST_{i % 40:02d}")
        ws.cell(row=row, column=15, value=_SECTIONS[i % len(_SECTIONS)])
        ws.cell(row=row, column=16, value=_LOCALITES[i % len(_LOCALITES)])
        ws.cell(row=row, column=17, value=_fake_phone(rng))
        ws.cell(row=row, column=18, value=f"PIECE_TEST_{i:06d}")
        ws.cell(row=row, column=19, value=_PIECE_NATURES[i % len(_PIECE_NATURES)])
        # Grille arbitraire, sans rapport avec une parcelle reelle.
        ws.cell(row=row, column=20, value=round(6.0 + (i % 100) * 0.001, 6))
        ws.cell(row=row, column=21, value=round(-6.0 - (i % 100) * 0.001, 6))
        row += 1


def _build_plantations_sheet(wb: Workbook, rng: random.Random) -> int:
    ws = wb.create_sheet("Registre plantations")
    ws.cell(row=1, column=1, value=f"COOP_DEMO — REGISTRE DES PLANTATIONS {CAMPAIGN}")
    _write_headers(ws, PLANTATION_HEADERS, 5)

    row = 8
    written = 0
    for i in range(1, N_PRODUCERS + 1):
        code = producer_code(i)
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=f"{code}-P1")
        ws.cell(row=row, column=3, value=_PROJETS[i % len(_PROJETS)])
        ws.cell(row=row, column=4, value=round(1.0 + (i % 50) * 0.1, 2))
        ws.cell(row=row, column=5, value=round(6.0 + (i % 100) * 0.001, 6))
        ws.cell(row=row, column=6, value=round(-6.0 - (i % 100) * 0.001, 6))
        ws.cell(row=row, column=7, value=450 + (i % 120))
        ws.cell(row=row, column=8, value=900 + (i % 300))
        row += 1
        written += 1

        # Seconde parcelle pour une partie des producteurs. Le code producteur
        # est VOLONTAIREMENT laisse vide : c'est le cas « cellule fusionnee »
        # que le parseur doit resoudre par propagation de la ligne precedente.
        if i <= N_EXTRA_PLANTATIONS:
            ws.cell(row=row, column=2, value=f"{code}-P2")
            ws.cell(row=row, column=3, value=_PROJETS[i % len(_PROJETS)])
            ws.cell(row=row, column=4, value=round(0.5 + (i % 20) * 0.1, 2))
            ws.cell(row=row, column=5, value=round(6.5 + (i % 100) * 0.001, 6))
            ws.cell(row=row, column=6, value=round(-6.5 - (i % 100) * 0.001, 6))
            row += 1
            written += 1
    return written


def _build_formations_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Registre Formation et Sensibili")
    ws.cell(row=1, column=1, value=f"COOP_DEMO — FORMATIONS {CAMPAIGN}")
    # Le parseur cherche, dans les 12 premieres lignes, une ligne dont la
    # colonne A contient « code » ET « producteur ».
    ws.cell(row=5, column=1, value="CODE PRODUCTEURS")
    for col, theme in enumerate(_THEMES, 2):
        ws.cell(row=5, column=col, value=theme)

    row = 6
    for i in range(1, N_FORMATION_ROWS + 1):
        ws.cell(row=row, column=1, value=producer_code(i))
        for col in range(2, 2 + len(_THEMES)):
            if (i + col) % 3 == 0:
                ws.cell(row=row, column=col, value="X")
        row += 1


def build_large_registry(target_dir: Path | str, filename: str = FILENAME) -> Path:
    """Ecrit le registre synthetique dans `target_dir` et renvoie son chemin.

    Deterministe : meme seed, meme contenu — un test qui echoue est
    reproductible a l'identique.
    """
    rng = random.Random(20260827)
    target = Path(target_dir) / filename

    wb = Workbook()
    wb.remove(wb.active)                    # retire la feuille par defaut
    _build_producers_sheet(wb, rng)
    _build_plantations_sheet(wb, rng)
    _build_formations_sheet(wb)
    wb.save(target)
    return target


if __name__ == "__main__":  # pragma: no cover - utilitaire manuel
    import sys
    out = build_large_registry(sys.argv[1] if len(sys.argv) > 1 else ".")
    print(f"OK -> {out}")
