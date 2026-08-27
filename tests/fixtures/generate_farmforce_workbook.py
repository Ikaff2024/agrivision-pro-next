"""Generateur de classeur FarmForce SYNTHETIQUE (format Fairtrade).

Remplace le classeur client reel qui servait de fixture. Reproduit la STRUCTURE
attendue par ``app.importers.farmforce_excel`` — noms de feuilles, adresses de
cellules, plages de lignes — avec des valeurs entierement fictives.

Le classeur d'origine etait un gabarit VIDE : le test se contentait donc de
verifier que le parseur ne plantait pas et que le total de revenus valait 0.
Cette fixture va plus loin et RENSEIGNE chaque feuille, ce qui permet de
verifier que le parseur extrait effectivement :

  * 1.profil               -> cooperative, producteur, localite, parcelles, menage
  * 2.entrees              -> revenus mensuels cacao et cafe, autres revenus
  * 3.couts                -> postes de couts
  * 4.main d'oeuvre        -> jours familiaux et main d'oeuvre salariee
  * 5.depenses du menage   -> depenses par categorie
  * 6.resultats            -> agregats de synthese
  * consent signatures     -> tracabilite des consentements

Aucune valeur ne provient d'un fichier client : noms explicitement de test,
cooperative fictive, montants arbitraires.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

# Cooperative fictive, sans rapport avec un client.
COOPERATIVE_NAME = "COOP_TEST_FARMFORCE"
PRODUCER_NAME = "PRODUCTEUR_TEST_FF001"
PRODUCER_CODE = "FFTEST-0001"
PR_CODE = "PR-TEST-0001"
LOCALITE = "LOCALITE_TEST_FF"
CAMPAIGN = "2025-2026"
FILENAME = f"digital data capturing tool DEMO SYNTHETIQUE {CAMPAIGN}.xlsx"

# Revenus cacao : 3 mois renseignes, quantite x prix unitaire.
CACAO_MONTHS = [("Octobre", 500, 1500), ("Novembre", 800, 1500), ("Decembre", 300, 1600)]
CAFE_MONTHS = [("Janvier", 120, 900)]
TOTAL_CACAO_REVENUE = sum(q * p for _, q, p in CACAO_MONTHS)
TOTAL_CAFE_REVENUE = sum(q * p for _, q, p in CAFE_MONTHS)
OTHER_REVENUE = 250_000
TOTAL_REVENUE = TOTAL_CACAO_REVENUE + TOTAL_CAFE_REVENUE + OTHER_REVENUE

COST_ITEMS = [("Engrais TEST", 120_000), ("Fongicide TEST", 45_000), ("Transport TEST", 30_000)]
TOTAL_COST = sum(amount for _, amount in COST_ITEMS)

LABOR_MONTHS = [("Octobre", 12, 8, 2500), ("Novembre", 15, 10, 2500)]
HOUSEHOLD_EXPENSES = [
    ("alimentation", 4, "Mensuel", 60_000),
    ("education", 11, "Trimestriel", 90_000),
    ("sante", 18, "Annuel", 40_000),
    ("autres", 25, "Mensuel", 25_000),
]
TOTAL_HOUSEHOLD_EXPENSES = sum(a for _, _, _, a in HOUSEHOLD_EXPENSES)


def _profil(wb: Workbook) -> None:
    ws = wb.create_sheet("1.profil")
    ws["D3"] = COOPERATIVE_NAME
    ws["D4"] = PR_CODE
    ws["D5"] = PRODUCER_NAME
    ws["H5"] = PRODUCER_CODE
    ws["D6"] = LOCALITE

    # Parcelles : lignes 14-23, colonnes B/C/E/G/I.
    for offset, (crop, surface, mode, age) in enumerate(
        [("Cacao", 2.5, "Plein soleil", 12), ("Cacao", 1.5, "Agroforestier", 7)]
    ):
        row = 14 + offset
        ws[f"B{row}"] = offset + 1
        ws[f"C{row}"] = crop
        ws[f"E{row}"] = surface
        ws[f"G{row}"] = mode
        ws[f"I{row}"] = age

    # Membres du menage travaillant sur la ferme : lignes 35-41.
    for offset, (name, rel, age, gender, job, pct) in enumerate(
        [("MEMBRE_TEST_A", "Conjoint", 40, "F", "Agriculture", 80),
         ("MEMBRE_TEST_B", "Enfant majeur", 22, "M", "Agriculture", 50)]
    ):
        row = 35 + offset
        ws[f"C{row}"] = name
        ws[f"D{row}"] = rel
        ws[f"E{row}"] = age
        ws[f"F{row}"] = gender
        ws[f"G{row}"] = job
        ws[f"I{row}"] = pct

    # Membres ne travaillant pas sur la ferme : lignes 45-60.
    ws["C45"] = "MEMBRE_TEST_C"
    ws["D45"] = "Enfant"
    ws["E45"] = 9
    ws["F45"] = "F"
    ws["G45"] = "Eleve"
    ws["H45"] = ""          # ne contribue pas au revenu


def _entrees(wb: Workbook) -> None:
    ws = wb.create_sheet("2.entrees")
    # Cacao : lignes 4-9, colonnes D (mois) / E (quantite) / F (prix).
    for offset, (month, qty, price) in enumerate(CACAO_MONTHS):
        row = 4 + offset
        ws[f"D{row}"] = month
        ws[f"E{row}"] = qty
        ws[f"F{row}"] = price
    # Cafe : lignes 20-25, memes colonnes.
    for offset, (month, qty, price) in enumerate(CAFE_MONTHS):
        row = 20 + offset
        ws[f"D{row}"] = month
        ws[f"E{row}"] = qty
        ws[f"F{row}"] = price
    # Autres revenus : ligne 35, produit en B, montant en K.
    ws["B35"] = "Vivriers TEST"
    ws["K35"] = OTHER_REVENUE


def _couts(wb: Workbook) -> None:
    ws = wb.create_sheet("3.couts")
    for offset, (product, cost) in enumerate(COST_ITEMS):
        row = 4 + offset
        ws[f"B{row}"] = product
        ws[f"E{row}"] = cost
        ws[f"G{row}"] = "Oui"


def _main_doeuvre(wb: Workbook) -> None:
    ws = wb.create_sheet("4.main d'oeuvre")
    # Le parseur balaie range(5, 39, 3) : une ligne sur trois.
    for offset, (month, family_days, hired_days, wage) in enumerate(LABOR_MONTHS):
        row = 5 + offset * 3
        ws[f"B{row}"] = month
        ws[f"C{row}"] = family_days
        ws[f"E{row}"] = hired_days
        ws[f"G{row}"] = wage
        ws[f"K{row}"] = hired_days * wage


def _depenses(wb: Workbook) -> None:
    ws = wb.create_sheet("5.depenses du menage")
    for _, row, period, amount in HOUSEHOLD_EXPENSES:
        ws[f"B{row}"] = period
        ws[f"C{row}"] = amount


def _resultats(wb: Workbook) -> None:
    ws = wb.create_sheet("6.resultats")
    ws["D10"] = TOTAL_REVENUE
    ws["E10"] = TOTAL_CACAO_REVENUE
    ws["D17"] = TOTAL_COST
    ws["E17"] = TOTAL_COST
    ws["D19"] = TOTAL_REVENUE - TOTAL_COST
    ws["E19"] = TOTAL_CACAO_REVENUE - TOTAL_COST
    ws["D23"] = sum(d for _, d, _, _ in LABOR_MONTHS)
    ws["E23"] = sum(d for _, d, _, _ in LABOR_MONTHS)
    ws["D24"] = 3_500
    ws["D30"] = TOTAL_HOUSEHOLD_EXPENSES


def _consent(wb: Workbook) -> None:
    ws = wb.create_sheet("consent signatures")
    ws["A2"] = "date"
    ws["B2"] = "Nom du producteur/trice"
    ws["C2"] = "consent to release to:"
    # Donnees a partir de la ligne 4 (cf. parseur).
    ws["A4"] = "2026-01-15"
    ws["B4"] = PRODUCER_NAME
    ws["C4"] = "X"
    ws["D4"] = "X"
    ws["E4"] = ""
    ws["F4"] = "X"


def build_farmforce_workbook(target_dir: Path | str, filename: str = FILENAME) -> Path:
    """Ecrit le classeur FarmForce synthetique et renvoie son chemin."""
    target = Path(target_dir) / filename
    wb = Workbook()
    wb.remove(wb.active)
    _profil(wb)
    _entrees(wb)
    _couts(wb)
    _main_doeuvre(wb)
    _depenses(wb)
    _resultats(wb)
    _consent(wb)
    wb.save(target)
    return target


if __name__ == "__main__":  # pragma: no cover - utilitaire manuel
    import sys
    out = build_farmforce_workbook(sys.argv[1] if len(sys.argv) > 1 else ".")
    print(f"OK -> {out}")
