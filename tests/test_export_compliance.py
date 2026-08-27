"""Blocage export EUDR + dérogation admin + rôle gestionnaire + composition Excel.

Règle métier : la production d'une parcelle EUDR non conforme ne peut pas être
EXPÉDIÉE (export_out), sauf dérogation expresse d'un administrateur (tracée).
La constitution du lot reste possible (le cacao peut entrer en magasin).
"""

from io import BytesIO

from openpyxl import load_workbook

from tests.conftest import create_member_headers


def _login(client, email, password="pass1234", role="admin", coop="Coop Export"):
    client.post("/auth/register", json={
        "email": email, "password": password, "role": role,
        "cooperative_name": coop, "country": "CI",
    })
    tok = client.post("/auth/login", json={"email": email, "password": password}).json()["access_token"]
    return {"Authorization": "Bearer " + tok}


def _plantation(client, h, name="P-NC", owner="Yao Export"):
    # Sans polygone / GPS / inspection -> EUDR non conforme (score 0/5).
    return client.post("/plantations", json={
        "name": name, "owner_name": owner, "country": "Côte d'Ivoire",
        "region": "Soubré", "hectares": 3.0,
    }, headers=h).json()


def _harvest(client, h, plantation_id, qty=500.0):
    return client.post(f"/plantations/{plantation_id}/harvests", json={
        "harvest_date": "2026-01-15T00:00:00", "quantity_kg": qty, "quality": "Bonne",
    }, headers=h).json()


def test_export_blocked_for_non_compliant_parcel_until_admin_waiver(client):
    """Expédition refusée (409) sur parcelle non conforme, autorisée après dérogation admin."""
    h = _login(client, "exp.block@test.ci", coop="Coop ExpBlock")
    p = _plantation(client, h)
    hv = _harvest(client, h, p["id"])
    lot = client.post("/lots", json={"harvest_ids": [hv["id"]]}, headers=h).json()

    # La constitution du lot reste possible ; l'EXPÉDITION est bloquée.
    r = client.post(f"/lots/{lot['id']}/movements", json={"movement_type": "export_out"}, headers=h)
    assert r.status_code == 409, r.text
    detail = r.json()["detail"]
    assert "non_compliant_plantations" in detail
    assert detail["non_compliant_plantations"][0]["plantation_id"] == p["id"]

    # Dérogation refusée à un non-admin (403)
    h_tech = create_member_headers(client, h, "exp.tech@test.ci", "technician")
    rt = client.post(f"/plantations/{p['id']}/export-waiver",
                     json={"reason": "tentative non autorisee"}, headers=h_tech)
    assert rt.status_code == 403

    # Dérogation admin -> expédition OK, dérogation tracée dans le mouvement
    rw = client.post(f"/plantations/{p['id']}/export-waiver",
                     json={"reason": "Plan de mise en conformite signe, echeance 30/09/2026"},
                     headers=h)
    assert rw.status_code == 200, rw.text
    assert rw.json()["export_waiver"] is True

    r2 = client.post(f"/lots/{lot['id']}/movements", json={"movement_type": "export_out"}, headers=h)
    assert r2.status_code == 200, r2.text
    assert r2.json()["status"] == "shipped"
    mv = [m for m in r2.json()["movements"] if m["movement_type"] == "export_out"][-1]
    assert mv["metadata"]["export_waivers"][0]["plantation_id"] == p["id"]


def test_export_waiver_revocation_blocks_again(client):
    """La révocation de la dérogation rebloque l'expédition ; statut visible dans eudr-status."""
    h = _login(client, "exp.rev@test.ci", coop="Coop ExpRev")
    p = _plantation(client, h)
    hv = _harvest(client, h, p["id"])
    lot = client.post("/lots", json={"harvest_ids": [hv["id"]]}, headers=h).json()

    client.post(f"/plantations/{p['id']}/export-waiver",
                json={"reason": "Derogation provisoire de test"}, headers=h)
    st = client.get(f"/plantations/{p['id']}/eudr-status", headers=h).json()
    assert st["export_waiver"] is True

    rd = client.delete(f"/plantations/{p['id']}/export-waiver", headers=h)
    assert rd.status_code == 200
    st2 = client.get(f"/plantations/{p['id']}/eudr-status", headers=h).json()
    assert st2["export_waiver"] is False

    r = client.post(f"/lots/{lot['id']}/movements", json={"movement_type": "export_out"}, headers=h)
    assert r.status_code == 409


def test_gestionnaire_role_operations(client):
    """Le gestionnaire fait les opérations administratives, sans les pouvoirs d'admin."""
    h_admin = _login(client, "exp.gadmin@test.ci", coop="Coop Gest")
    h_g = create_member_headers(client, h_admin, "gestionnaire@test.ci", "gestionnaire")

    # Autorisé : plantation, récolte, entrepôt, lot
    p = client.post("/plantations", json={
        "name": "P-Gest", "owner_name": "Akissi G", "country": "CI", "hectares": 2.0,
    }, headers=h_g)
    assert p.status_code in (200, 201), p.text
    pj = p.json()
    hv = client.post(f"/plantations/{pj['id']}/harvests", json={
        "harvest_date": "2026-01-15T00:00:00", "quantity_kg": 250, "quality": "Bonne",
    }, headers=h_g)
    assert hv.status_code == 201, hv.text
    w = client.post("/warehouses", json={"name": "Magasin G"}, headers=h_g)
    assert w.status_code == 201, w.text
    lot = client.post("/lots", json={"harvest_ids": [hv.json()["id"]]}, headers=h_g)
    assert lot.status_code == 201, lot.text

    # Interdit : création de membre, dérogation export (réservés admin)
    m = client.post("/admin/members", json={"email": "x@test.ci", "role": "technician"}, headers=h_g)
    assert m.status_code == 403
    wv = client.post(f"/plantations/{pj['id']}/export-waiver",
                     json={"reason": "pas autorise du tout"}, headers=h_g)
    assert wv.status_code == 403


def test_lot_composition_xlsx_format_exporter(client):
    """L'export Excel de composition reprend les colonnes du fichier exportateur."""
    h = _login(client, "exp.xlsx@test.ci", coop="Coop Xlsx")
    p = _plantation(client, h, name="P-X", owner="Brou X")
    hv = _harvest(client, h, p["id"], qty=220.0)
    lot = client.post("/lots", json={
        "harvest_ids": [hv["id"]],
        "exporter": "OCEAN-SA",
        "external_ref": "Lot N°41/SC100035-2025-2-121",
    }, headers=h).json()
    assert lot["exporter"] == "OCEAN-SA"

    r = client.get(f"/lots/{lot['id']}/composition.xlsx", headers=h)
    assert r.status_code == 200, r.text
    assert "Composition_" in r.headers.get("content-disposition", "")

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == (
        "Cooperative Name", "Export lot N°/Connaissement", "Date of purchase from cooperative",
        "Certification", "Farmer_ID", "Farm_ID", "Net Weight (KG)", "Exporter",
    )
    assert len(rows) == 2
    body = rows[1]
    assert body[0] == "Coop Xlsx"
    assert body[1] == "Lot N°41/SC100035-2025-2-121"
    assert body[6] == 220.0
    assert body[7] == "OCEAN-SA"
    assert body[5].endswith("-P1")  # Farm_ID = <code producteur>-P<rang>

    # PATCH des infos export du lot
    pr = client.patch(f"/lots/{lot['id']}", json={"exporter": "CARGILL"}, headers=h)
    assert pr.status_code == 200 and pr.json()["exporter"] == "CARGILL"
