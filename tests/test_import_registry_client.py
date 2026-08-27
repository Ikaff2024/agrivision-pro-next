"""Import de registre cooperative — parseur eprouve sur un registre SYNTHETIQUE.

Ces tests utilisaient auparavant un registre client reel versionne dans le
depot. Le fichier est desormais genere a la volee par
``tests/fixtures/generate_large_registry.py`` : meme structure, meme echelle,
aucune donnee personnelle. L'intention couverte est inchangee — volume,
detection de campagne, mapping des colonnes, conventions de codes — et deux
proprietes que l'ancien test ne verifiait pas explicitement sont desormais
assertees : la resolution des colonnes ambigues et la propagation des cellules
fusionnees.
"""
import pytest

from app.importers.cooperative_registry import parse_registry
from tests.fixtures.generate_large_registry import (
    CODE_PREFIX,
    N_EXTRA_PLANTATIONS,
    N_PRODUCERS,
    build_large_registry,
    producer_code,
)


@pytest.fixture(scope="module")
def large_registry(tmp_path_factory):
    """Registre synthetique volumineux, genere une seule fois par session.

    Genere hors du depot (tmp_path de pytest) : aucun binaire de plusieurs Mo
    n'est versionne.
    """
    target = tmp_path_factory.mktemp("registry")
    return build_large_registry(target)


@pytest.fixture(scope="module")
def parsed_large(large_registry):
    return parse_registry(str(large_registry), filename=large_registry.name)


# ── Volume et integrite globale ─────────────────────────────────────────────

def test_large_registry_parses_without_error(parsed_large):
    assert parsed_large.errors == []


def test_large_registry_detects_campaign_from_filename(parsed_large):
    assert parsed_large.detected_campaign == "2025-2026"


def test_large_registry_keeps_full_scale(parsed_large):
    """Le parseur doit rester eprouve a l'echelle d'une grosse cooperative.

    Seuils repris de l'ancien test sur fichier client (>6900 producteurs,
    >7000 parcelles) pour que la portee du test soit inchangee.
    """
    assert len(parsed_large.producers) > 6900
    assert len(parsed_large.plantations) > 7000
    assert len(parsed_large.producers) == N_PRODUCERS
    assert len(parsed_large.plantations) == N_PRODUCERS + N_EXTRA_PLANTATIONS


def test_large_registry_summary_exposes_formations(parsed_large):
    summary = parsed_large.summary()
    assert "formations" in summary
    assert summary["formations"] > 0


# ── Conventions de codes ────────────────────────────────────────────────────

def test_producer_and_plantation_codes_follow_convention(parsed_large):
    first = parsed_large.producers[0]
    assert first.code_yeyasso == producer_code(1)
    assert first.code_yeyasso.startswith(CODE_PREFIX)
    assert parsed_large.plantations[0].code_plantation == f"{producer_code(1)}-P1"


# ── Mapping des colonnes ────────────────────────────────────────────────────

def test_producer_columns_are_mapped(parsed_large):
    """Chaque colonne detectee par mot-cle doit atterrir dans le bon champ."""
    first = parsed_large.producers[0]
    assert first.nom_complet == "PRODUCTEUR_TEST_00001"
    assert first.code_plantation == f"{producer_code(1)}-P1"
    assert first.section.startswith("SECTION_TEST_")
    assert first.localite.startswith("LOCALITE_TEST_")
    assert first.formateur_interne_nom.startswith("FORMATEUR_TEST_")
    assert first.piece_numero == "PIECE_TEST_000001"
    assert first.piece_nature.endswith("_TEST")
    assert first.telephone, "le contact doit etre extrait"
    assert first.latitude is not None and first.longitude is not None


def test_ambiguous_columns_resolve_to_producer_not_household(parsed_large):
    """« sexe » et « date de naissance » apparaissent DEUX fois dans le registre.

    Le premier bloc decrit un membre du menage, le second le producteur. Le
    parseur doit retenir la DERNIERE occurrence. La fixture met des valeurs
    differentes dans les deux blocs, donc une inversion est immediatement
    visible ici — ce que l'ancien test ne verifiait pas.
    """
    first = parsed_large.producers[0]
    assert first.sexe == "M", "sexe du MENAGE ('F') retenu au lieu de celui du producteur"
    assert first.date_naissance is not None
    assert first.date_naissance.year == 1975, (
        "date de naissance du MENAGE (1990) retenue au lieu de celle du producteur"
    )


def test_merged_producer_code_is_propagated_to_second_plantation(parsed_large):
    """Cellules fusionnees : la 2e parcelle d'un producteur n'a pas son code.

    Le parseur doit propager le dernier code vu. Sans propagation, ces
    parcelles seraient orphelines.
    """
    seconds = [p for p in parsed_large.plantations if p.code_plantation.endswith("-P2")]
    assert len(seconds) == N_EXTRA_PLANTATIONS
    for plantation in seconds[:20]:
        expected = plantation.code_plantation.rsplit("-P", 1)[0]
        assert plantation.code_producteur == expected, (
            "code producteur non propage depuis la cellule fusionnee"
        )


def test_plantation_attributes_are_mapped(parsed_large):
    first = parsed_large.plantations[0]
    assert first.superficie_ha and first.superficie_ha > 0
    assert first.latitude is not None and first.longitude is not None
    assert first.projet in {"FT-RA", "FT", "RA", "EUDR"}


# ── Registre sans feuille producteurs ───────────────────────────────────────

def test_plantation_only_registry_synthesizes_producers(tmp_path):
    """Registre centre plantations : les producteurs sont deduits des codes.

    Reprend le cas « Coop test 2 » sans dependre d'un fichier versionne.
    """
    from openpyxl import Workbook

    path = tmp_path / "Registre DEMO plantations seules 2025-2026.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Registre plantations"
    for col, label in enumerate(
        ["CODE PRODUCTEUR", "CODE PLANTATION", "PROJET", "Superficie (ha)"], 1
    ):
        ws.cell(row=5, column=col, value=label)
    row = 8
    for i in range(1, 151):
        code = f"PRODTEST{i:04d}"
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=f"{code}-P1")
        ws.cell(row=row, column=3, value="FT-RA")
        ws.cell(row=row, column=4, value=2.0)
        row += 1
    wb.save(path)

    result = parse_registry(str(path), filename=path.name)

    assert result.errors == []
    assert len(result.plantations) > 100
    # Un producteur deduit par code distinct.
    assert len(result.producers) == len(result.plantations)
    assert result.producers[0].code_yeyasso == "PRODTEST0001"
    # Nom provisoire = code, en attendant un vrai registre producteurs.
    assert result.producers[0].nom_complet == "PRODTEST0001"
    # La certification portee par la parcelle est conservee.
    assert result.producers[0].projet == "FT-RA"
