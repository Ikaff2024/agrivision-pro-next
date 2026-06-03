"""
Génère un petit registre coopérative .xlsx valide pour les tests E2E d'import.

Format minimal accepté par app/importers/cooperative_registry.py :
- une feuille nommée « Plantations » (hint de détection) ;
- en-têtes en ligne 5 (les en-têtes peuvent s'étaler sur les lignes 5-7) ;
- données à partir de la ligne 7 (on commence en ligne 8, la 7 reste vide) ;
- colonnes détectées par mots-clés : « code producteur », « code plantation »,
  « superficie », « latitude », « longitude ».

Pas de feuille « producteurs » : ils sont synthétisés depuis les plantations.

Régénérer :  python e2e/fixtures/make_registry.py
"""
import os
from openpyxl import Workbook

OUT = os.path.join(os.path.dirname(__file__), "registre_demo_e2e.xlsx")

HEADERS = [
    "CODE PRODUCTEUR", "CODE PLANTATION", "PROJET",
    "SUPERFICIE (HA)", "LATITUDE", "LONGITUDE",
]
ROWS = [
    ("E2E-001", "E2E-001-P1", "EUDR", 2.5, 6.10, -6.80),
    ("E2E-002", "E2E-002-P1", "EUDR", 1.8, 6.12, -6.82),
    ("E2E-003", "E2E-003-P1", "EUDR", 3.1, 6.14, -6.84),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantations"
    for col, label in enumerate(HEADERS, 1):
        ws.cell(row=5, column=col, value=label)  # en-têtes ligne 5
    r = 8  # données à partir de la ligne 8 (la 7 reste vide)
    for row in ROWS:
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)
        r += 1
    wb.save(OUT)
    print(f"OK -> {OUT} ({len(ROWS)} plantations)")


if __name__ == "__main__":
    main()
