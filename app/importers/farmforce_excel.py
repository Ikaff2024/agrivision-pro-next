from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional


@dataclass
class FarmForceParseResult:
    producer_code: Optional[str] = None
    producer_name: Optional[str] = None
    cooperative_name: Optional[str] = None
    campaign_label: str = "2025-2026"
    localite: Optional[str] = None
    pr_code: Optional[str] = None
    household_members: list[dict] = field(default_factory=list)
    parcels: list[dict] = field(default_factory=list)
    revenue_items: list[dict] = field(default_factory=list)
    cost_items: list[dict] = field(default_factory=list)
    family_labor_items: list[dict] = field(default_factory=list)
    hired_labor_items: list[dict] = field(default_factory=list)
    food_security_items: list[dict] = field(default_factory=list)
    household_expenses: list[dict] = field(default_factory=list)
    consent_records: list[dict] = field(default_factory=list)
    summary: dict = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def as_payload(self) -> dict:
        return {
            "producer_id": None,
            "campaign_label": self.campaign_label,
            "localite": self.localite,
            "pr_code": self.pr_code or self.producer_code,
            "household_members": self.household_members,
            "parcels": self.parcels,
            "revenue_items": self.revenue_items,
            "cost_items": self.cost_items,
            "family_labor_items": self.family_labor_items,
            "hired_labor_items": self.hired_labor_items,
            "food_security_items": self.food_security_items,
            "notes": "Import Excel FarmForce Fairtrade",
        }


def _value(ws, cell: str) -> Any:
    return ws[cell].value


def _text(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _num(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        if isinstance(value, str):
            value = value.replace(",", ".").replace(" ", "").strip()
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _nonzero(*values: Any) -> bool:
    return any(v not in (None, "", 0, "0") for v in values)


def _append_monthly_revenue(result: FarmForceParseResult, product: str, unit: str, rows: list[tuple[str, Any, Any]]):
    for month, quantity, price in rows:
        quantity_n = _num(quantity)
        price_n = _num(price)
        if quantity_n or price_n:
            result.revenue_items.append({
                "month": month,
                "product": product,
                "unit": unit,
                "quantity": quantity_n,
                "unit_price_cfa": price_n,
                "revenue_cfa": quantity_n * price_n,
            })


def parse_farmforce_excel(file_path: str, filename: str | None = None) -> FarmForceParseResult:
    from openpyxl import load_workbook

    result = FarmForceParseResult()
    if filename and "2025-2026" in filename:
        result.campaign_label = "2025-2026"

    try:
        wb = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as exc:
        result.errors.append(f"Impossible d'ouvrir le fichier FarmForce: {exc}")
        return result

    try:
        if "1.profil" in wb.sheetnames:
            ws = wb["1.profil"]
            result.cooperative_name = _text(_value(ws, "D3"))
            result.producer_name = _text(_value(ws, "D5"))
            result.producer_code = _text(_value(ws, "H5"))
            result.localite = _text(_value(ws, "D6"))
            result.pr_code = _text(_value(ws, "D4"))

            for row in range(14, 24):
                crop = _text(_value(ws, f"C{row}"))
                surface = _num(_value(ws, f"E{row}"))
                management = _text(_value(ws, f"G{row}"))
                age = _num(_value(ws, f"I{row}"))
                if _nonzero(crop, surface, management, age):
                    result.parcels.append({
                        "parcel": str(_value(ws, f"B{row}") or row - 13),
                        "crop": crop,
                        "surface_ha": surface,
                        "management_mode": management,
                        "age_years": age,
                    })

            for row in range(35, 42):
                name = _text(_value(ws, f"C{row}"))
                if name:
                    result.household_members.append({
                        "name": name,
                        "relationship": _text(_value(ws, f"D{row}")),
                        "age": _num(_value(ws, f"E{row}")),
                        "gender": _text(_value(ws, f"F{row}")),
                        "occupation": _text(_value(ws, f"G{row}")),
                        "works_on_farm": True,
                        "agricultural_time_pct": _num(_value(ws, f"I{row}")),
                    })
            for row in range(45, 61):
                name = _text(_value(ws, f"C{row}"))
                if name:
                    result.household_members.append({
                        "name": name,
                        "relationship": _text(_value(ws, f"D{row}")),
                        "age": _num(_value(ws, f"E{row}")),
                        "gender": _text(_value(ws, f"F{row}")),
                        "occupation": _text(_value(ws, f"G{row}")),
                        "works_on_farm": False,
                        "contributes_income": bool(_text(_value(ws, f"H{row}"))),
                    })

        if "2.entrees" in wb.sheetnames:
            ws = wb["2.entrees"]
            cacao_rows = [(str(_value(ws, f"D{r}") or ""), _value(ws, f"E{r}"), _value(ws, f"F{r}")) for r in range(4, 10)]
            cacao_rows += [(str(_value(ws, f"H{r}") or ""), _value(ws, f"I{r}"), _value(ws, f"J{r}")) for r in range(4, 10)]
            _append_monthly_revenue(result, "Cacao", "kg", cacao_rows)

            cafe_rows = [(str(_value(ws, f"D{r}") or ""), _value(ws, f"E{r}"), _value(ws, f"F{r}")) for r in range(20, 26)]
            cafe_rows += [(str(_value(ws, f"H{r}") or ""), _value(ws, f"I{r}"), _value(ws, f"J{r}")) for r in range(20, 26)]
            _append_monthly_revenue(result, "Cafe", "kg", cafe_rows)

            for row in list(range(35, 43)) + list(range(47, 57)) + list(range(87, 95)):
                product = _text(_value(ws, f"B{row}")) or _text(_value(ws, f"G{row}"))
                revenue = _num(_value(ws, f"K{row}") or _value(ws, f"F{row}"))
                if product and revenue:
                    result.revenue_items.append({
                        "product": product,
                        "quantity": 1,
                        "unit_price_cfa": revenue,
                        "revenue_cfa": revenue,
                    })

        if "3.couts" in wb.sheetnames:
            ws = wb["3.couts"]
            for row in list(range(4, 14)) + list(range(19, 32)) + list(range(37, 47)) + list(range(52, 62)):
                product = _text(_value(ws, f"B{row}")) or _text(_value(ws, f"C{row}"))
                cost = _num(_value(ws, f"E{row}") or _value(ws, f"F{row}"))
                if product and cost:
                    result.cost_items.append({
                        "category": "FarmForce",
                        "product": product,
                        "cost_cfa": cost,
                        "used_for_cocoa": _text(_value(ws, f"G{row}")),
                    })

        if "4.main d'oeuvre" in wb.sheetnames:
            ws = wb["4.main d'oeuvre"]
            for row in range(5, 39, 3):
                month = _text(_value(ws, f"B{row}"))
                if not month:
                    continue
                family_days = _num(_value(ws, f"C{row}"))
                hired_days = _num(_value(ws, f"E{row}"))
                wage = _num(_value(ws, f"G{row}"))
                if family_days:
                    result.family_labor_items.append({"month": month, "total_days": family_days})
                if hired_days or wage:
                    result.hired_labor_items.append({
                        "month": month,
                        "total_days": hired_days,
                        "daily_wage_cfa": wage,
                        "labor_cost_cfa": _num(_value(ws, f"K{row}")) or hired_days * wage,
                    })

        if "5.depenses du menage" in wb.sheetnames:
            ws = wb["5.depenses du menage"]
            for label, rows in {
                "alimentation": range(4, 8),
                "education": range(11, 15),
                "sante": range(18, 22),
                "autres": range(25, 29),
            }.items():
                for row in rows:
                    amount = _num(_value(ws, f"C{row}"))
                    if amount:
                        result.household_expenses.append({
                            "category": label,
                            "period": _text(_value(ws, f"B{row}")),
                            "cost_cfa": amount,
                        })

        if "6.resultats" in wb.sheetnames:
            ws = wb["6.resultats"]
            result.summary = {
                "total_revenue_cfa": _num(_value(ws, "D10")),
                "total_cocoa_revenue_cfa": _num(_value(ws, "E10")),
                "total_cost_cfa": _num(_value(ws, "D17")),
                "total_cocoa_cost_cfa": _num(_value(ws, "E17")),
                "profit_cfa": _num(_value(ws, "D19")),
                "cocoa_profit_cfa": _num(_value(ws, "E19")),
                "family_labor_days": _num(_value(ws, "D23")),
                "cocoa_family_labor_days": _num(_value(ws, "E23")),
                "return_on_labor_cfa": _num(_value(ws, "D24")),
                "household_expenses_cfa": _num(_value(ws, "D30")),
            }

        if "consent signatures" in wb.sheetnames:
            ws = wb["consent signatures"]
            for row in range(4, ws.max_row + 1):
                name = _text(_value(ws, f"B{row}"))
                if name:
                    result.consent_records.append({
                        "date": str(_value(ws, f"A{row}") or ""),
                        "producer_name": name,
                        "fairtrade_international": bool(_text(_value(ws, f"C{row}"))),
                        "fairtrade_africa": bool(_text(_value(ws, f"D{row}"))),
                        "spo": bool(_text(_value(ws, f"E{row}"))),
                        "buyers": bool(_text(_value(ws, f"F{row}"))),
                    })

        if not result.producer_code and not result.producer_name:
            result.warnings.append("Aucun producteur renseigne dans l'onglet 1.profil.")

    finally:
        wb.close()

    return result
