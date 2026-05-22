"""
cooperative_registry.py
=========================
Parser du fichier Excel "Registre coopérative" (type YEYASSO).

Lit les 3 feuilles (Producteurs, Plantations, Formations) et retourne
des dataclasses Python. Ne touche PAS la base de donnees : l'insertion
est faite par la couche d'import (Phase 0.1.b-2).

Usage :
    from app.importers.cooperative_registry import parse_registry
    result = parse_registry("/chemin/Registre_FT_YEYASSO.xlsx")
    print(result.summary())
    for prod in result.producers:
        ...
"""
from __future__ import annotations
from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Optional, Any
import logging
import unicodedata

logger = logging.getLogger("agrivision.importer")

# Extensions image acceptees (sans objet, juste pour reference)
SHEET_PRODUCERS_HINTS = ["producteur", "membre"]
SHEET_PLANTATIONS_HINTS = ["plantation"]
SHEET_FORMATIONS_HINTS = ["formation", "sensibili"]


# ===========================================================================
# Dataclasses de sortie
# ===========================================================================

@dataclass
class ParsedProducer:
    code_yeyasso: str
    nom_complet: str
    code_plantation: Optional[str] = None
    code_saco: Optional[str] = None
    recepisse: Optional[str] = None
    projet: Optional[str] = None              # ex "FT-RA" -> certifications
    sexe: Optional[str] = None
    date_naissance: Optional[date] = None
    formateur_interne_nom: Optional[str] = None
    section: Optional[str] = None
    localite: Optional[str] = None
    telephone: Optional[str] = None
    piece_numero: Optional[str] = None
    piece_nature: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    row_index: int = 0                        # ligne Excel d'origine (debug)


@dataclass
class ParsedDelivery:
    date_vente: Optional[date]
    poids_kg: Optional[float]
    nbre_sacs: Optional[int]
    numero_recu: Optional[str]


@dataclass
class ParsedPlantation:
    code_plantation: str
    code_producteur: Optional[str] = None
    projet: Optional[str] = None
    superficie_ha: Optional[float] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    production_estimee_kg: Optional[float] = None
    rendement_kg_ha: Optional[float] = None
    deliveries: list = field(default_factory=list)
    row_index: int = 0


@dataclass
class ParseWarning:
    sheet: str
    row_index: int
    message: str


@dataclass
class RegistryParseResult:
    producers: list = field(default_factory=list)
    plantations: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    detected_campaign: Optional[str] = None
    errors: list = field(default_factory=list)

    def summary(self) -> dict:
        return {
            "producers": len(self.producers),
            "plantations": len(self.plantations),
            "deliveries": sum(len(p.deliveries) for p in self.plantations),
            "warnings": len(self.warnings),
            "errors": len(self.errors),
            "detected_campaign": self.detected_campaign,
        }


# ===========================================================================
# Helpers de normalisation
# ===========================================================================

def _strip_accents(text: str) -> str:
    """Retire les accents pour comparaison robuste."""
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _norm(text: Any) -> str:
    """Normalise une chaine : minuscule, sans accent, espaces compresses."""
    if text is None:
        return ""
    s = _strip_accents(str(text)).lower()
    return " ".join(s.split())


def parse_float(value: Any) -> Optional[float]:
    """Convertit en float, tolerant. Retourne None si impossible."""
    if value is None or value == "":
        return None
    try:
        if isinstance(value, str):
            value = value.replace(",", ".").replace(" ", "").strip()
        return float(value)
    except (ValueError, TypeError):
        return None


def parse_int(value: Any) -> Optional[int]:
    f = parse_float(value)
    return int(f) if f is not None else None


def parse_date_any(value: Any) -> Optional[date]:
    """
    Convertit une date en objet date. Gere :
    - datetime/date natifs (openpyxl les renvoie pour les cellules date)
    - chaines ISO "1981-10-17 00:00:00"
    - chaines FR "28/01/1975"
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    # ISO
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def normalize_phone(value: Any) -> Optional[str]:
    """Normalise un numero de telephone (chiffres uniquement)."""
    if value is None or value == "":
        return None
    digits = "".join(c for c in str(value) if c.isdigit())
    return digits or None


# ===========================================================================
# Recherche floue de colonnes
# ===========================================================================

def find_column(headers: dict, *keyword_groups) -> Optional[int]:
    """
    Cherche l'index de colonne dont l'en-tete contient TOUS les mots-cles
    d'au moins un groupe.

    headers : dict {col_index: "texte en-tete normalise"}
    keyword_groups : chaque argument est une liste de mots-cles.
        find_column(h, ["code", "yeya"]) -> colonne contenant "code" ET "yeya"
        find_column(h, ["code", "yeya"], ["code", "membre"]) -> l'un OU l'autre

    Retourne le premier index correspondant, ou None.
    """
    for group in keyword_groups:
        for col_idx in sorted(headers.keys()):
            if all(kw in headers[col_idx] for kw in group):
                return col_idx
    return None


def find_columns_all(headers: dict, keywords: list) -> list:
    """
    Retourne TOUS les index de colonnes dont l'en-tete contient tous les
    mots-cles, tries par position croissante. Utile quand un meme libelle
    apparait plusieurs fois (ex : 'sexe' du menage ET 'sexe' du producteur).
    """
    return [
        col_idx for col_idx in sorted(headers.keys())
        if all(kw in headers[col_idx] for kw in keywords)
    ]


def build_headers(worksheet, header_rows: list) -> dict:
    """
    Construit le dict {col_index: en-tete_normalise} en fusionnant
    plusieurs lignes d'en-tete (ex L5 + L6).
    Optimise pour les worksheets read_only : une seule passe iter_rows.
    """
    headers = {}
    max_row_needed = max(header_rows)
    rows_data = {}
    row_idx = 0
    for row in worksheet.iter_rows(min_row=1, max_row=max_row_needed, values_only=True):
        row_idx += 1
        if row_idx in header_rows:
            rows_data[row_idx] = row

    # Determiner le nombre de colonnes
    max_col = max((len(r) for r in rows_data.values()), default=0)
    for col in range(1, max_col + 1):
        parts = []
        for hr in header_rows:
            row = rows_data.get(hr)
            if row and col - 1 < len(row):
                val = row[col - 1]
                if val is not None:
                    parts.append(_norm(val))
        headers[col] = " ".join(p for p in parts if p)
    return headers


# ===========================================================================
# Parsing de la feuille Producteurs
# ===========================================================================

def _parse_producers_sheet(ws, result: RegistryParseResult):
    """Parse la feuille Registre Producteurs (optimise iter_rows)."""
    # En-tetes sur lignes 5 et 6, donnees a partir de la ligne 7
    headers = build_headers(ws, [5, 6])

    col_code   = find_column(headers, ["code", "yeya"])
    col_nom    = find_column(headers, ["nom", "membre"], ["nom", "prenom"])
    col_plant  = find_column(headers, ["code", "plantation"])
    col_saco   = find_column(headers, ["code", "saco"])
    col_recep  = find_column(headers, ["recepisse"])
    col_projet = find_column(headers, ["projet"])
    col_format = find_column(headers, ["formateur"])
    col_section = find_column(headers, ["section"])
    col_localite = find_column(headers, ["localite"])
    col_contact = find_column(headers, ["contact"])
    col_piece_num = find_column(headers, ["piece", "identification"], ["numero", "piece"])
    col_piece_nat = find_column(headers, ["nature", "piece"])
    col_lon = find_column(headers, ["longitude"])
    col_lat = find_column(headers, ["latitude"])

    # Cas particulier : "sexe" et "date de naissance" apparaissent DEUX fois
    # (une fois pour le membre du menage, une fois pour le producteur).
    # La colonne du PRODUCTEUR est la derniere occurrence (apres la zone menage).
    sexe_cols = find_columns_all(headers, ["sexe"])
    col_sexe = sexe_cols[-1] if sexe_cols else None

    dnaiss_cols = find_columns_all(headers, ["date", "naissance"])
    # La date de naissance du producteur est la derniere occurrence.
    col_dnaiss = dnaiss_cols[-1] if dnaiss_cols else None

    if col_code is None or col_nom is None:
        result.errors.append(
            "Feuille Producteurs : colonnes 'code' ou 'nom' introuvables. "
            "Verifiez que les en-tetes sont bien sur les lignes 5-6."
        )
        return

    def get(row_tuple, col_idx):
        """Recupere une valeur depuis le tuple iter_rows (col 1-indexed)."""
        if col_idx is None:
            return None
        i = col_idx - 1
        return row_tuple[i] if 0 <= i < len(row_tuple) else None

    # iter_rows : lecture sequentielle rapide. min_row=7 = debut donnees.
    row_idx = 6
    for row in ws.iter_rows(min_row=7, values_only=True):
        row_idx += 1

        code = get(row, col_code)
        nom = get(row, col_nom)

        # Ligne vide ou ligne de total -> on ignore silencieusement
        if not code or not nom:
            continue

        code_s = str(code).strip()
        nom_s = str(nom).strip().upper()

        if not code_s or not any(c.isalnum() for c in code_s):
            result.warnings.append(ParseWarning(
                "Producteurs", row_idx, f"Code invalide ignore : '{code_s}'"
            ))
            continue

        prod = ParsedProducer(
            code_yeyasso=code_s,
            nom_complet=nom_s,
            code_plantation=str(get(row, col_plant)).strip() if get(row, col_plant) else None,
            code_saco=str(get(row, col_saco)).strip() if get(row, col_saco) else None,
            recepisse=str(get(row, col_recep)).strip() if get(row, col_recep) else None,
            projet=str(get(row, col_projet)).strip() if get(row, col_projet) else None,
            sexe=str(get(row, col_sexe)).strip().upper()[:1] if get(row, col_sexe) else None,
            date_naissance=parse_date_any(get(row, col_dnaiss)),
            formateur_interne_nom=str(get(row, col_format)).strip() if get(row, col_format) else None,
            section=str(get(row, col_section)).strip() if get(row, col_section) else None,
            localite=str(get(row, col_localite)).strip() if get(row, col_localite) else None,
            telephone=normalize_phone(get(row, col_contact)),
            piece_numero=str(get(row, col_piece_num)).strip() if get(row, col_piece_num) else None,
            piece_nature=str(get(row, col_piece_nat)).strip() if get(row, col_piece_nat) else None,
            latitude=parse_float(get(row, col_lat)),
            longitude=parse_float(get(row, col_lon)),
            row_index=row_idx,
        )
        result.producers.append(prod)

        if prod.telephone is None:
            result.warnings.append(ParseWarning(
                "Producteurs", row_idx, f"{nom_s} : pas de telephone"
            ))


# ===========================================================================
# Parsing de la feuille Plantations
# ===========================================================================

def _parse_plantations_sheet(ws, result: RegistryParseResult):
    """
    Parse la feuille Registre Plantations (optimise iter_rows).
    Les livraisons sont sur les colonnes 21+ (4 colonnes par livraison :
    Date / Poids / Nb sacs / N recu).
    """
    headers = build_headers(ws, [5, 6, 7])

    col_code_prod = find_column(headers, ["code", "producteur"])
    col_code_plant = find_column(headers, ["code", "plantation"])
    col_projet = find_column(headers, ["projet"])
    col_superf = find_column(headers, ["superficie"])
    col_lon = find_column(headers, ["longitude"])
    col_lat = find_column(headers, ["latitude"])
    col_prod_est = find_column(headers, ["production", "estimee"])
    col_rendement = find_column(headers, ["rendement", "2025"])

    if col_code_plant is None:
        result.errors.append(
            "Feuille Plantations : colonne 'code plantation' introuvable."
        )
        return

    def get(row_tuple, col_idx):
        if col_idx is None:
            return None
        i = col_idx - 1
        return row_tuple[i] if 0 <= i < len(row_tuple) else None

    row_idx = 6
    for row in ws.iter_rows(min_row=7, values_only=True):
        row_idx += 1

        code_plant = get(row, col_code_plant)
        if not code_plant:
            continue

        plant = ParsedPlantation(
            code_plantation=str(code_plant).strip(),
            code_producteur=str(get(row, col_code_prod)).strip() if get(row, col_code_prod) else None,
            projet=str(get(row, col_projet)).strip() if get(row, col_projet) else None,
            superficie_ha=parse_float(get(row, col_superf)),
            latitude=parse_float(get(row, col_lat)),
            longitude=parse_float(get(row, col_lon)),
            production_estimee_kg=parse_float(get(row, col_prod_est)),
            rendement_kg_ha=parse_float(get(row, col_rendement)),
            row_index=row_idx,
        )

        # Livraisons : balayage des colonnes 21+ par groupes de 4 (0-indexed: 20+)
        n = len(row)
        for start0 in range(20, n, 4):
            d = row[start0] if start0 < n else None
            poids = row[start0 + 1] if start0 + 1 < n else None
            sacs = row[start0 + 2] if start0 + 2 < n else None
            recu = row[start0 + 3] if start0 + 3 < n else None

            date_v = parse_date_any(d)
            poids_v = parse_float(poids)

            if date_v and poids_v:
                plant.deliveries.append(ParsedDelivery(
                    date_vente=date_v,
                    poids_kg=poids_v,
                    nbre_sacs=parse_int(sacs),
                    numero_recu=str(recu).strip() if recu else None,
                ))
            elif (date_v and not poids_v) or (poids_v and not date_v):
                result.warnings.append(ParseWarning(
                    "Plantations", row_idx,
                    f"{plant.code_plantation} : livraison incomplete ignoree (col {start0 + 1})"
                ))

        result.plantations.append(plant)


# ===========================================================================
# Detection de la campagne
# ===========================================================================

def _detect_campaign(workbook, filename: str) -> Optional[str]:
    """
    Detecte la campagne agricole depuis le nom de fichier ou le contenu.
    Cherche un motif 4 chiffres consecutifs (annee) ou AAAA-AAAA / AAAA_AAAA.
    """
    import re
    # 1. Depuis le nom de fichier
    m = re.search(r"(20\d{2})[_-]?(20\d{2})", filename or "")
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r"(20\d{2})", filename or "")
    if m:
        year = int(m.group(1))
        return f"{year}-{year + 1}"
    # 2. Depuis les en-tetes des feuilles
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in range(1, 8):
            for col in range(1, min((ws.max_column or 0) + 1, 130)):
                val = ws.cell(row=row, column=col).value
                if val:
                    mm = re.search(r"(20\d{2})[_/ -](20\d{2})", str(val))
                    if mm:
                        return f"{mm.group(1)}-{mm.group(2)}"
    return None


# ===========================================================================
# Point d'entree principal
# ===========================================================================

def find_sheet(workbook, hints: list) -> Optional[str]:
    """Trouve une feuille dont le nom contient l'un des indices."""
    for name in workbook.sheetnames:
        norm = _norm(name)
        if any(h in norm for h in hints):
            return name
    return None


def parse_registry(file_path: str, filename: str = None) -> RegistryParseResult:
    """
    Parse un fichier Registre cooperative complet.

    Args:
        file_path : chemin du fichier .xlsx sur le disque
        filename  : nom d'origine du fichier (pour detecter la campagne)

    Returns:
        RegistryParseResult avec producers, plantations, warnings, errors.
    """
    from openpyxl import load_workbook

    result = RegistryParseResult()

    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        result.errors.append(f"Impossible d'ouvrir le fichier : {e}")
        return result

    try:
        result.detected_campaign = _detect_campaign(wb, filename or file_path)

        sheet_prod = find_sheet(wb, SHEET_PRODUCERS_HINTS)
        sheet_plant = find_sheet(wb, SHEET_PLANTATIONS_HINTS)

        if sheet_prod:
            logger.info("Parsing feuille producteurs : %s", sheet_prod)
            _parse_producers_sheet(wb[sheet_prod], result)
        else:
            result.errors.append("Aucune feuille 'producteurs' detectee.")

        if sheet_plant:
            logger.info("Parsing feuille plantations : %s", sheet_plant)
            _parse_plantations_sheet(wb[sheet_plant], result)
        else:
            result.warnings.append(ParseWarning(
                "global", 0, "Aucune feuille 'plantations' detectee."
            ))

    finally:
        wb.close()

    logger.info("Parsing termine : %s", result.summary())
    return result
